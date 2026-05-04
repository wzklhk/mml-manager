#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MML配置文件转Excel/CSV脚本
支持命令行参数指定输入文件
默认生成Excel，每个命令类型一个Sheet页
可选生成CSV
"""
import re
import os
import argparse
import csv
from typing import List, Dict, Set
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Error: openpyxl is required. Install it using 'pip install openpyxl'")
    exit(1)

from utils.parse import parse_any_command
from utils.sort import sort_configs


def convert_file_to_excel_and_csv(input_file: str, output_base: str, generate_excel: bool = True, generate_csv: bool = False, encoding: str = 'utf-8'):
    """
    转换配置文件为Excel和/或CSV

    Args:
        input_file: 输入文件路径
        output_base: 输出文件基础路径（不含扩展名）
        generate_excel: 是否生成Excel
        generate_csv: 是否生成CSV
        encoding: 文件编码
    """
    print(f"正在读取文件: {input_file}")

    with open(input_file, 'r', encoding=encoding) as f:
        lines = f.readlines()

    # 统计数据 - 支持所有MML命令类型
    configs_by_table = {}  # table_name -> list of config dicts
    all_columns = {}  # table_name -> set of columns

    for line_num, line in enumerate(lines, 1):
        line = line.strip()
        if not line or line.startswith('--'):
            continue

        if line.startswith('ENTER'):
            continue

        elif line.startswith('SET') or line.startswith('ADD'):
            config = parse_any_command(line)
            if config:
                table_name = config['table']
                if table_name not in configs_by_table:
                    configs_by_table[table_name] = []
                    all_columns[table_name] = set()

                configs_by_table[table_name].append(config)
                all_columns[table_name].update(config['values'].keys())

    # 打印解析统计
    print(f"\n解析完成:")
    total_configs = 0
    for table in sorted(configs_by_table.keys()):
        count = len(configs_by_table[table])
        total_configs += count
        print(f"  - {table:<30} : {count:>4} 条")
    print(f"  - 总计: {total_configs} 条配置")

    # 对每个表的记录进行升序排序
    for table_name in configs_by_table:
        configs_by_table[table_name] = sort_configs(configs_by_table[table_name])

    print(f"\n已对每个表的记录进行升序排序")

    # 生成Excel
    if generate_excel:
        excel_path = f"{output_base}.xlsx"
        print(f"\n正在生成Excel文件: {excel_path}")
        wb = Workbook()
        # 删除默认创建的sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        for table_name in sorted(all_columns.keys()):
            ws = wb.create_sheet(title=table_name)
            
            # 准备列头和数据
            columns = sorted(list(all_columns[table_name]))
            
            # 写入表头
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 写入数据
            for row_idx, config in enumerate(configs_by_table[table_name], 2):
                values = config['values']
                for col_idx, col_name in enumerate(columns, 1):
                    value = values.get(col_name, '')
                    # Try to convert numeric strings to actual numbers for Excel storage
                    if value != '':
                        try:
                            # Try integer first
                            if '.' not in str(value):
                                value = int(value)
                            else:
                                value = float(value)
                        except (ValueError, TypeError):
                            # Keep as string if conversion fails
                            pass
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 自动调整列宽 (简单估算)
            for col_idx, col_name in enumerate(columns, 1):
                max_len = len(str(col_name))
                # 检查前几行数据以估算宽度，避免全量扫描影响性能
                for r in range(2, min(len(configs_by_table[table_name]) + 2, 20)):
                    cell_val = ws.cell(row=r, column=col_idx).value
                    if cell_val:
                        max_len = max(max_len, len(str(cell_val)))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

        wb.save(excel_path)
        print(f"[OK] Excel文件已生成: {excel_path}")

    # 生成CSV
    if generate_csv:
        csv_dir = f"{output_base}_csv"
        os.makedirs(csv_dir, exist_ok=True)
        print(f"\n正在生成CSV文件到目录: {csv_dir}")
        
        for table_name in sorted(all_columns.keys()):
            csv_path = os.path.join(csv_dir, f"{table_name}.csv")
            columns = sorted(list(all_columns[table_name]))
            
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=columns)
                writer.writeheader()
                
                for config in configs_by_table[table_name]:
                    # Ensure all columns are present, fill missing with empty string
                    row_data = {col: config['values'].get(col, '') for col in columns}
                    writer.writerow(row_data)
                    
        print(f"[OK] CSV文件已生成: {csv_dir}")


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='MML配置文件转Excel/CSV脚本',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 默认只生成Excel
  python mml2xls.py input.txt

  # 只生成CSV
  python mml2xls.py input.txt --csv

  # 同时生成Excel和CSV
  python mml2xls.py input.txt --both

  # 指定输出文件名前缀
  python mml2xls.py input.txt -o my_output
        """
    )

    parser.add_argument('input_file', nargs='?',
                        help='输入的MML配置文件路径')
    parser.add_argument('-o', '--output',
                        help='输出的文件基础路径（不含扩展名，默认：同名）')
    
    output_group = parser.add_mutually_exclusive_group()
    output_group.add_argument('--excel', action='store_true', default=True,
                        help='生成Excel文件 (默认)')
    output_group.add_argument('--csv', action='store_true',
                        help='只生成CSV文件')
    output_group.add_argument('--both', action='store_true',
                        help='同时生成Excel和CSV文件')
                        
    parser.add_argument('--encoding', default='utf-8',
                        help='文件编码（默认：utf-8）')

    args = parser.parse_args()

    # 如果没有提供输入文件，显示帮助
    if not args.input_file:
        parser.print_help()
        return

    # 检查输入文件是否存在
    if not os.path.exists(args.input_file):
        print(f"❌ 错误: 输入文件不存在!")
        print(f"   路径: {args.input_file}")
        return

    # 确定输出路径
    input_dir = os.path.dirname(os.path.abspath(args.input_file))
    input_name = os.path.splitext(os.path.basename(args.input_file))[0]

    output_base = args.output if args.output else os.path.join(input_dir, input_name)
    
    generate_excel = False
    generate_csv = False
    
    if args.csv:
        generate_csv = True
    elif args.both:
        generate_excel = True
        generate_csv = True
    else:
        # Default is excel
        generate_excel = True

    print("=" * 60)
    print("MML配置文件转Excel/CSV脚本")
    print("=" * 60)
    print(f"\n输入文件: {args.input_file}")
    if generate_excel:
        print(f"输出Excel: {output_base}.xlsx")
    if generate_csv:
        print(f"输出CSV目录: {output_base}_csv/")
    print()

    # 转换文件
    try:
        convert_file_to_excel_and_csv(
            args.input_file, 
            output_base, 
            generate_excel=generate_excel, 
            generate_csv=generate_csv,
            encoding=args.encoding
        )

        print("\n" + "=" * 60)
        print("转换完成!")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n[ERROR] 错误: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()