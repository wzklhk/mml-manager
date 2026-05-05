#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel/CSV文件转MML配置命令脚本

支持将Excel文件（.xlsx）或CSV文件转换回MML配置命令格式。
每个Sheet页（Excel）或每个CSV文件会被转换为对应表名的SET命令。

用法:
  # Excel 转 MML（默认SET命令）
  python xls2mml.py input.xlsx

  # 指定输出文件
  python xls2mml.py input.xlsx -o output.txt

  # 使用ADD命令
  python xls2mml.py input.xlsx --cmd ADD

  # CSV 转 MML
  python xls2mml.py data.csv

  # 目录下所有CSV转MML
  python xls2mml.py ./csv_dir --csv
"""
import os
import argparse
from typing import List, Dict, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install it using 'pip install openpyxl'")
    exit(1)


def quote_mml_value(value) -> str:
    """
    将值格式化为MML值
    - None -> ''
    - 数字 -> 直接转为字符串
    - 字符串 -> 如果包含空格、逗号、分号或双引号，加双引号并转义
    """
    if value is None:
        return ''
    if isinstance(value, float):
        # 将浮点整数（如 201.0）转为整数字符串
        if value == int(value):
            return str(int(value))
        return str(value)
    value = str(value).strip()
    if not value:
        return ''
    # 如果值包含空格、逗号、分号或双引号，需要加引号
    import re
    if re.search(r'[\s,;"]', value):
        escaped = value.replace('"', '""')
        return f'"{escaped}"'
    return value


def excel_to_mml(input_path: str, output_path: str, cmd_type: str = 'SET', sheet_names: Optional[List[str]] = None):
    """
    将Excel文件(.xlsx)转换为MML配置命令

    Args:
        input_path: 输入Excel文件路径
        output_path: 输出MML文件路径
        cmd_type: MML命令类型 (SET/ADD)
        sheet_names: 要处理的Sheet页名称列表，None表示全部处理
    """
    print(f"正在读取Excel文件: {input_path}")
    wb = load_workbook(input_path, read_only=True, data_only=True)

    if sheet_names:
        sheets = [s for s in sheet_names if s in wb.sheetnames]
        if not sheets:
            print(f"⚠️ 警告: 未找到指定的Sheet页: {sheet_names}")
            print(f"   可用Sheet页: {wb.sheetnames}")
            return
    else:
        sheets = wb.sheetnames

    total_lines = 0
    mml_lines = []

    for sheet_name in sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # 第一行是列头
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        # 过滤掉空列头
        valid_indices = [i for i, h in enumerate(headers) if h]
        valid_headers = [headers[i] for i in valid_indices]

        if not valid_headers:
            print(f"⚠️ 跳过Sheet '{sheet_name}': 无有效列头")
            continue

        # 确定表名：如果Sheet名包含特殊字符则不作为表名
        import re
        table_name = re.sub(r'[^a-zA-Z0-9_]', '_', sheet_name)
        if not table_name or table_name[0].isdigit():
            table_name = f"T_{table_name}"

        sheet_count = 0
        for row_idx, row in enumerate(rows[1:], 2):
            # 跳过全空行
            values = {}
            has_value = False
            for i in valid_indices:
                val = row[i] if i < len(row) else None
                val_str = quote_mml_value(val)
                if val_str:
                    has_value = True
                values[valid_headers[i]] = val_str

            if not has_value:
                continue

            # 构建MML命令: SET 表名:KEY1=val1,KEY2=val2;
            kv_pairs = []
            for key in valid_headers:
                val = values.get(key, '')
                if val != '':
                    kv_pairs.append(f"{key}={val}")

            if kv_pairs:
                mml_line = f"{cmd_type} {table_name}:{','.join(kv_pairs)};"
                # 添加注释标明数据来源行
                mml_lines.append(mml_line)
                mml_lines.append(f"-- 来源: Sheet='{sheet_name}', 行={row_idx}")
                sheet_count += 1

        print(f"  - {sheet_name:<30} : {sheet_count:>4} 条")
        total_lines += sheet_count

    wb.close()

    print(f"\n总计: {total_lines} 条配置命令")

    # 写入输出文件
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"-- 由 xls2mml.py 从 {os.path.basename(input_path)} 生成\n")
        f.write(f"-- 生成时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"-- 命令类型: {cmd_type}\n\n")
        for line in mml_lines:
            f.write(line + '\n')

    print(f"[OK] MML文件已生成: {output_path}")


def csv_to_mml(input_path: str, output_path: str, cmd_type: str = 'SET', table_name: Optional[str] = None):
    """
    将CSV文件转换为MML配置命令

    Args:
        input_path: 输入CSV文件路径
        output_path: 输出MML文件路径
        cmd_type: MML命令类型 (SET/ADD)
        table_name: 表名，默认使用CSV文件名（不含扩展名）
    """
    import csv

    print(f"正在读取CSV文件: {input_path}")

    if table_name is None:
        import re
        table_name = os.path.splitext(os.path.basename(input_path))[0]
        table_name = re.sub(r'[^a-zA-Z0-9_]', '_', table_name)
        if not table_name or table_name[0].isdigit():
            table_name = f"T_{table_name}"

    mml_lines = []
    count = 0

    with open(input_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        if not headers:
            print(f"⚠️ 未找到列头，跳过: {input_path}")
            return

        valid_headers = [h.strip() for h in headers if h.strip()]

        for row_idx, row in enumerate(reader, 2):
            kv_pairs = []
            has_value = False

            for key in valid_headers:
                val = row.get(key, '').strip() if row.get(key) else ''
                if val:
                    has_value = True
                val_str = quote_mml_value(val)
                if val_str:
                    kv_pairs.append(f"{key}={val_str}")

            if not has_value:
                continue

            if kv_pairs:
                mml_line = f"{cmd_type} {table_name}:{','.join(kv_pairs)};"
                mml_lines.append(mml_line)
                mml_lines.append(f"-- 来源: CSV行={row_idx}")
                count += 1

    print(f"  - {table_name:<30} : {count:>4} 条")
    print(f"总计: {count} 条配置命令")

    # 写入输出文件
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"-- 由 xls2mml.py 从 {os.path.basename(input_path)} 生成\n")
        f.write(f"-- 生成时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"-- 命令类型: {cmd_type}\n")
        f.write(f"-- 表名: {table_name}\n\n")
        for line in mml_lines:
            f.write(line + '\n')

    print(f"[OK] MML文件已生成: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description='Excel/CSV文件转MML配置命令脚本',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # Excel 转 MML（默认SET命令）
  python xls2mml.py input.xlsx

  # 使用ADD命令
  python xls2mml.py input.xlsx --cmd ADD

  # CSV 转 MML
  python xls2mml.py data.csv -o output.mml

  # 指定表名（CSV文件默认用文件名做表名）
  python xls2mml.py data.csv --table MY_TABLE
        """
    )

    parser.add_argument('input_file',
                        help='输入的Excel(.xlsx)或CSV文件路径')
    parser.add_argument('-o', '--output',
                        help='输出的MML文件路径（默认：同名.mml）')
    parser.add_argument('--cmd', default='SET', choices=['SET', 'ADD'],
                        help='MML命令类型，默认: SET')
    parser.add_argument('--table',
                        help='表名（CSV模式有效，默认使用CSV文件名）')
    parser.add_argument('--sheets', nargs='*',
                        help='Excel中要处理的Sheet页名称（不指定则处理所有）')

    args = parser.parse_args()

    # 检查输入文件是否存在
    if not os.path.exists(args.input_file):
        print(f"❌ 错误: 输入文件不存在!")
        print(f"   路径: {args.input_file}")
        return

    # 确定输出路径
    input_dir = os.path.dirname(os.path.abspath(args.input_file))
    input_name = os.path.splitext(os.path.basename(args.input_file))[0]
    output_path = args.output if args.output else os.path.join(input_dir, f"{input_name}.mml")

    print("=" * 60)
    print("Excel/CSV转MML配置命令脚本")
    print("=" * 60)
    print(f"输入文件: {args.input_file}")
    print(f"输出文件: {output_path}")
    print(f"命令类型: {args.cmd}")
    print()

    ext = os.path.splitext(args.input_file)[1].lower()

    if ext in ('.csv',):
        csv_to_mml(args.input_file, output_path, cmd_type=args.cmd, table_name=args.table)
    elif ext in ('.xlsx', '.xls'):
        excel_to_mml(args.input_file, output_path, cmd_type=args.cmd, sheet_names=args.sheets)
    else:
        print(f"⚠️ 不支持的文件格式: {ext}")
        print(f"   支持: .xlsx, .xls, .csv")
        return

    print("\n" + "=" * 60)
    print("转换完成!")
    print("=" * 60)


if __name__ == '__main__':
    main()
