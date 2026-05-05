import openpyxl
import sys
import os


def convert_xls_to_mml(input_file, output_file=None):
    """
    Converts an Excel file (.xlsx) where each sheet represents MML commands.
    
    Assumptions:
    - Row 1 contains headers: [Command_Name, Param1, Param2, ...]
    - Subsequent rows contain data.
    - MML Format generated: CMD_NAME(Param1=Value1, Param2=Value2);
    
    :param input_file: Path to the input .xlsx file
    :param output_file: Path to the output .mml or .txt file. If None, prints to stdout.
    """
    if not os.path.exists(input_file):
        print(f"Error: File {input_file} not found.")
        return

    try:
        workbook = openpyxl.load_workbook(input_file)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    mml_output_lines = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        mml_output_lines.append(f"-- Sheet: {sheet_name} --")

        # Get all rows
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            continue

        # Assume first row is header
        headers = rows[0]
        
        # Filter out empty headers to determine valid parameter columns
        # Index 0 is assumed to be the Command Name
        cmd_name_idx = 0
        param_headers = []
        param_indices = []
        
        for idx, header in enumerate(headers):
            if idx == 0:
                continue
            if header is not None and str(header).strip() != "":
                param_headers.append(str(header).strip())
                param_indices.append(idx)

        # Process data rows
        for row in rows[1:]:
            # Skip empty rows
            if not any(cell is not None for cell in row):
                continue

            cmd_name = row[cmd_name_idx]
            if cmd_name is None or str(cmd_name).strip() == "":
                continue
            
            cmd_name = str(cmd_name).strip()
            
            # Build parameters part
            params = []
            for i, p_idx in enumerate(param_indices):
                if p_idx < len(row):
                    val = row[p_idx]
                    if val is not None:
                        # Format value: quote strings, leave numbers as is
                        if isinstance(val, str):
                            val_str = f'"{val}"'
                        else:
                            val_str = str(val)
                        params.append(f"{param_headers[i]}={val_str}")
            
            # Construct MML command
            # Format: CMD_NAME(PARAM1=VAL1, PARAM2=VAL2);
            param_str = ", ".join(params)
            if param_str:
                mml_line = f"{cmd_name}({param_str});"
            else:
                mml_line = f"{cmd_name};"
                
            mml_output_lines.append(mml_line)
        
        mml_output_lines.append("")  # Empty line between sheets

    final_output = "\n".join(mml_output_lines)

    if output_file:
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(final_output)
            print(f"Successfully converted to {output_file}")
        except Exception as e:
            print(f"Error writing output file: {e}")
    else:
        print(final_output)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python xls2mml.py <input.xlsx> [output.mml]")
        sys.exit(1)

    input_excel = sys.argv[1]
    output_mml = sys.argv[2] if len(sys.argv) > 2 else None
    
    convert_xls_to_mml(input_excel, output_mml)