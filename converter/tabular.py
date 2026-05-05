# -*- coding: utf-8 -*-
"""
表格格式读写模块。

提供Excel (.xlsx) 和 CSV 文件的统一读写接口。
所有表格输入输出操作集中在此，便于扩展其他格式（如ODS）。
"""
import os
import csv
from typing import Dict, List, Optional, Tuple

from utils.table import MmlDataSet, TableGroup, MmlConfig
from utils.mml import quote_mml_value


# ============================================================
#  写入：MML数据集 → Excel/CSV
# ============================================================

def write_excel(file_path: str, dataset: MmlDataSet, styled: bool = True):
    """
    将MML数据集写入Excel文件，每个表一个Sheet。

    Args:
        file_path: 输出.xlsx路径
        dataset: MML数据集
        styled: 是否带表头样式
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    # 删除默认Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for table_name in dataset.tables:
        group = dataset.get_group(table_name)
        if not group or not group.configs:
            continue

        ws = wb.create_sheet(title=table_name)
        columns = sorted(group.columns)

        if styled:
            _write_excel_header(ws, columns)
        else:
            _write_excel_row(ws, 1, columns)

        for row_idx, config in enumerate(group.configs, 2 if styled else 2):
            values = config.values
            row_data = [_str_val(values.get(col, '')) for col in columns]
            _write_excel_row(ws, row_idx, row_data)

        # 自动列宽
        _auto_column_width(ws, columns, group.configs)

    wb.save(file_path)
    print(f"[OK] Excel文件已生成: {file_path}")


def write_csvs(output_dir: str, dataset: MmlDataSet, prefix: str = ''):
    """
    将MML数据集写入多个CSV文件，每个表一个文件。

    Args:
        output_dir: 输出目录
        dataset: MML数据集
        prefix: 输出目录前缀（默认基于output_dir）
    """
    os.makedirs(output_dir, exist_ok=True)

    for table_name in dataset.tables:
        group = dataset.get_group(table_name)
        if not group or not group.configs:
            continue

        csv_path = os.path.join(output_dir, f"{prefix}{table_name}.csv")
        columns = sorted(group.columns)

        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            for config in group.configs:
                row_data = {col: config.values.get(col, '') for col in columns}
                writer.writerow(row_data)

    print(f"[OK] CSV文件已生成到: {output_dir}")


# ============================================================
#  读取：Excel/CSV → MML数据集
# ============================================================

def read_excel(file_path: str, sheet_names: Optional[List[str]] = None,
               pick_fields: Optional[List[str]] = None,
               skip_empty_rows: bool = True) -> MmlDataSet:
    """
    从Excel文件读取数据为MML数据集。

    Args:
        file_path: 输入.xlsx路径
        sheet_names: 指定Sheet名（None=全部）
        pick_fields: 仅提取指定字段（None=全部）
        skip_empty_rows: 跳过全空行

    Returns:
        MmlDataSet
    """
    from openpyxl import load_workbook

    dataset = MmlDataSet()
    wb = load_workbook(file_path, read_only=True, data_only=True)

    sheets = _resolve_sheets(wb, sheet_names)
    if not sheets:
        wb.close()
        return dataset

    print(f"正在读取Excel文件: {file_path}")
    for sheet_name in sheets:
        ws = wb[sheet_name]
        table_name = sanitize_from_name(sheet_name)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers, valid_indices = _parse_headers(rows[0], pick_fields)
        if not headers:
            continue

        count = 0
        for row in rows[1:]:
            kv = _build_kv_from_row(row, headers, valid_indices, skip_empty_rows)
            if kv is not None:
                dataset.add(MmlConfig(cmd_type='SET', table=table_name, values=kv))
                count += 1

        print(f"  - {sheet_name:<30} : {count:>4} 条")

    wb.close()
    return dataset


def read_csv(file_path: str, table_name: Optional[str] = None,
             pick_fields: Optional[List[str]] = None) -> MmlDataSet:
    """
    从CSV文件读取数据为MML数据集。

    Args:
        file_path: 输入CSV路径
        table_name: 目标表名（默认使用文件名）
        pick_fields: 仅提取指定字段（None=全部）

    Returns:
        MmlDataSet
    """
    import re

    if table_name is None:
        table_name = re.sub(r'[^a-zA-Z0-9_]', '_',
                            os.path.splitext(os.path.basename(file_path))[0])
        if not table_name or table_name[0].isdigit():
            table_name = f"T_{table_name}"

    dataset = MmlDataSet()
    print(f"正在读取CSV文件: {file_path}")

    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        if not fieldnames:
            print(f"⚠️ 无有效列头: {file_path}")
            return dataset

        valid_headers = _filter_headers(fieldnames, pick_fields)
        count = 0

        for row in reader:
            kv = {}
            has_value = False
            for key in valid_headers:
                val = row.get(key, '').strip() if row.get(key) else ''
                if val:
                    has_value = True
                    kv[key] = quote_mml_value(val)
                else:
                    kv[key] = ''
            if not has_value:
                continue
            dataset.add(MmlConfig(cmd_type='SET', table=table_name, values=kv))
            count += 1

        print(f"  - {table_name:<30} : {count:>4} 条")

    return dataset


def read_csv_batch(directory: str, output_dir: str,
                   cmd_type: str = 'SET') -> MmlDataSet:
    """
    批量读取目录下所有CSV文件。

    Args:
        directory: CSV文件目录
        output_dir: MML输出目录（自动创建）
        cmd_type: 命令类型

    Returns:
        合并的数据集
    """
    os.makedirs(output_dir, exist_ok=True)

    csv_files = sorted([f for f in os.listdir(directory) if f.lower().endswith('.csv')])
    if not csv_files:
        print(f"⚠️ 目录 '{directory}' 中没有CSV文件")
        return MmlDataSet()

    print(f"找到 {len(csv_files)} 个CSV文件，开始批量转换...\n")
    combined = MmlDataSet()

    for i, csv_file in enumerate(csv_files, 1):
        input_path = os.path.join(directory, csv_file)
        ds = read_csv(input_path)
        for table in ds.tables:
            group = ds.get_group(table)
            if group:
                for c in group.configs:
                    c.cmd_type = cmd_type
                    combined.add(c)

        # 写入单个MML文件
        output_name = os.path.splitext(csv_file)[0] + '.mml'
        output_path = os.path.join(output_dir, output_name)
        from utils.io_handler import write_mml_file
        write_mml_file(output_path, ds,
                       source_desc=f"csv2mml.py 从 {csv_file}",
                       cmd_type=cmd_type)
        print()

    print(f"批量转换完成! 共处理 {len(csv_files)} 个CSV文件")
    return combined


# ============================================================
#  内部辅助函数
# ============================================================

def _str_val(v) -> str:
    """值转字符串，数值优化"""
    if v is None or v == '':
        return ''
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    return str(v)


def _write_excel_header(ws, columns):
    """写入带样式的Excel表头"""
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def _write_excel_row(ws, row_idx, data):
    """写入一行Excel数据"""
    for ci, val in enumerate(data, 1):
        ws.cell(row=row_idx, column=ci, value=val)


def _auto_column_width(ws, columns, configs):
    """自动列宽估算"""
    from openpyxl.utils import get_column_letter
    for ci, col in enumerate(columns, 1):
        max_len = len(str(col))
        for cfg in configs[:20]:
            val = cfg.values.get(col, '')
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 50)


def _resolve_sheets(wb, sheet_names):
    """解析Sheet名列表"""
    if sheet_names:
        sheets = [s for s in sheet_names if s in wb.sheetnames]
        if not sheets:
            print(f"⚠️ 未找到指定Sheet: {sheet_names}")
            print(f"   可用: {wb.sheetnames}")
        return sheets
    return wb.sheetnames


def _parse_headers(row, pick_fields):
    """解析表头行，返回(valid_headers, valid_indices)"""
    import re
    headers = [str(h).strip() if h is not None else '' for h in row]
    valid_indices = [i for i, h in enumerate(headers) if h]
    all_headers = [headers[i] for i in valid_indices]

    if pick_fields:
        filtered = [(i, h) for i, h in zip(valid_indices, all_headers) if h in pick_fields]
        return [h for _, h in filtered], [i for i, _ in filtered]

    return all_headers, valid_indices


def _build_kv_from_row(row, headers, indices, skip_empty):
    """从Excel行构建键值对"""
    kv = {}
    has_value = False
    for idx, key in zip(indices, headers):
        val = row[idx] if idx < len(row) else None
        qv = quote_mml_value(val)
        if qv:
            has_value = True
        kv[key] = qv
    return kv if (has_value or not skip_empty) else None


def _filter_headers(fieldnames, pick_fields):
    """过滤列头列表"""
    stripped = [h.strip() for h in fieldnames if h.strip()]
    if pick_fields:
        return [h for h in stripped if h in pick_fields]
    return stripped


def sanitize_from_name(name: str) -> str:
    """从Sheet名/文件名生成表名"""
    import re
    table = re.sub(r'[^a-zA-Z0-9_]', '_', name)
    if not table or table[0].isdigit():
        table = f"T_{table}"
    return table
