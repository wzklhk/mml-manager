# -*- coding: utf-8 -*-
"""MML → XLS/CSV 转换核心模块"""

import os
import csv
from typing import Dict, List, Set, Optional

from converters.mml_to_sql import parse_mml_file, sort_configs_by_values


def convert_file_to_excel_and_csv(
    input_file: str,
    output_base: str,
    generate_excel: bool = True,
    generate_csv: bool = False,
    encoding: str = "utf-8",
) -> Dict:
    """转换 MML 文件为 Excel 和/或 CSV。

    Returns:
        {excel_path, csv_dir, total, tables: {table_name: count}}
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise ImportError("需要 openpyxl: pip install openpyxl")

    configs_by_table, all_columns = parse_mml_file(input_file, encoding)

    result = {
        "total": sum(len(v) for v in configs_by_table.values()),
        "tables": {tn: len(v) for tn, v in configs_by_table.items()},
    }

    if generate_excel:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for table_name in sorted(all_columns.keys()):
            columns = sorted(list(all_columns[table_name]))
            ws = wb.create_sheet(title=table_name)

            # 表头
            for ci, col in enumerate(columns, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # 数据
            for ri, config in enumerate(configs_by_table[table_name], 2):
                for ci, col in enumerate(columns, 1):
                    val = config["values"].get(col, "")
                    ws.cell(row=ri, column=ci, value=val)

        excel_path = f"{output_base}.xlsx"
        wb.save(excel_path)
        result["excel_path"] = excel_path

    if generate_csv:
        csv_dir = f"{output_base}_csv"
        os.makedirs(csv_dir, exist_ok=True)

        for table_name in sorted(all_columns.keys()):
            columns = sorted(list(all_columns[table_name]))
            csv_path = os.path.join(csv_dir, f"{table_name}.csv")
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                for config in configs_by_table[table_name]:
                    writer.writerow([config["values"].get(col, "") for col in columns])

        result["csv_dir"] = csv_dir

    return result
