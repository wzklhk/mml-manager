# -*- coding: utf-8 -*-
"""XLS → MML 转换核心模块"""

import os
from typing import Dict, Optional

from utils.mml import format_mml_value_simple


def convert_xls_to_mml(input_file: str, output_file: str = None) -> Dict:
    """转换 Excel 文件为 MML 格式。

    格式：第一列为表名 + 命令类型，后续列为参数。
    如: SUBRACK SET  -> SET SUBRACK:ID=1,NAME=xxx;

    Returns:
        {total, output_path}
    """
    import openpyxl

    wb = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    mml_lines = []
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h else "" for h in rows[0]]

        for row in rows[1:]:
            row = [str(c).strip() if c else "" for c in row]
            if not any(row):
                continue

            # 第一列为表名:命令类型，如 "SUBRACK SET"
            cmd_cell = row[0]
            parts = cmd_cell.split()
            if len(parts) < 2:
                continue
            cmd_type = parts[-1]  # SET/ADD
            table_name = " ".join(parts[:-1])

            values = {}
            for i in range(1, min(len(headers), len(row))):
                if headers[i] and row[i]:
                    values[headers[i]] = format_mml_value_simple(row[i])

            if values:
                mml_parts = [f"{k}={v}" for k, v in values.items()]
                mml_lines.append(f"{cmd_type} {table_name}:{','.join(mml_parts)};\n")
                total += 1

    wb.close()

    if output_file is None:
        output_file = os.path.splitext(input_file)[0] + ".mml"

    header = f"-- 由 xls2mml.py 从 {os.path.basename(input_file)} 生成\n\n"
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(header)
        f.writelines(mml_lines)

    return {"total": total, "output_path": output_file}
